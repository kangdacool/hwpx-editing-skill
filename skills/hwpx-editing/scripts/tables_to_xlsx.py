#!/usr/bin/env python3
"""
tables_to_xlsx.py — export HWPX tables to Excel (.xlsx), preserving merged cells.

Usage:
    python tables_to_xlsx.py FILE.hwpx [OUT.xlsx]

Each top-level data table becomes a sheet (T1, T2, ...). Cell merges
(colSpan/rowSpan) are reproduced as Excel merged ranges. Single-cell layout
frames and nested tables are skipped. A cell whose only content is an image or
an equation is written as [그림] / [수식] (Excel can't hold those inline).

Note: extraction reflects the file's *actual* structure. If a table's merges were
already lost upstream (e.g. a lossy .hwp → .hwpx conversion), they can't be
recovered here — that information is already gone from the file.

Requires: lxml + openpyxl  (pip install openpyxl)
"""
import argparse
import os
import sys
import zipfile

# Force UTF-8 console output so 한글/글리프가 cp949 콘솔에서도 안 깨지게.
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError, OSError):
    pass

import hwpxlib as H

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError as e:  # pragma: no cover
    raise SystemExit(
        "tables_to_xlsx needs openpyxl. Install it with:  pip install openpyxl"
    ) from e

P = H.P  # "{http://www.hancom.co.kr/hwpml/2011/paragraph}"


def _cell_text(tc) -> str:
    """This cell's own text (excludes nested-table cell text). Falls back to a
    [그림]/[수식] marker when the cell holds only an image or equation."""
    parts = []
    for t in tc.findall(f".//{P}t"):
        if next((a for a in t.iterancestors() if a.tag == f"{P}tc"), None) is tc:
            parts.append("".join(t.itertext()))
    txt = " ".join(" ".join(parts).split())
    if not txt and tc.find(f".//{P}equation") is not None:
        txt = "[수식]"
    if not txt and tc.find(f".//{P}pic") is not None:
        txt = "[그림]"
    return txt


def convert(path: str, out: str) -> int:
    """Write every top-level data table in `path` to `out` (.xlsx), one sheet each,
    merged cells preserved. Returns the number of sheets written."""
    H.ensure_hwpx(path)  # legacy .hwp → clear guidance instead of a crash
    z = zipfile.ZipFile(path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    n = 0
    for name in H.section_names(z):
        root = H.etree.fromstring(z.read(name))
        for tbl in root.findall(f".//{P}tbl"):
            if next((a for a in tbl.iterancestors() if a.tag == f"{P}tbl"), None) is not None:
                continue  # nested table — its own sheet would be contextless
            rc = int(tbl.get("rowCnt") or 0)
            cc = int(tbl.get("colCnt") or 0)
            if rc <= 1 and cc <= 1:
                continue  # 1x1 layout frame, not a data table
            n += 1
            ws = wb.create_sheet(title=f"T{n}")
            for tc in tbl.findall(f".//{P}tc"):
                if next((a for a in tc.iterancestors() if a.tag == f"{P}tc"), None) is not None:
                    continue  # cell belonging to a nested table
                addr = tc.find(f"{P}cellAddr")
                span = tc.find(f"{P}cellSpan")
                c = int(addr.get("colAddr"))
                r = int(addr.get("rowAddr"))
                cs = int((span.get("colSpan") if span is not None else "1") or 1)
                rs = int((span.get("rowSpan") if span is not None else "1") or 1)
                ws.cell(row=r + 1, column=c + 1, value=_cell_text(tc))
                if cs > 1 or rs > 1:
                    ws.merge_cells(start_row=r + 1, start_column=c + 1,
                                   end_row=r + rs, end_column=c + cs)
            for col in range(1, cc + 1):
                ws.column_dimensions[get_column_letter(col)].width = 16
    if n == 0:
        raise SystemExit("No data tables found (only 1x1 layout frames / nested tables).")
    wb.save(out)
    return n


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Export HWPX tables to Excel (.xlsx), merges preserved.")
    ap.add_argument("file", help="the .hwpx to read")
    ap.add_argument("out", nargs="?",
                    help="output .xlsx (default: <file>_tables.xlsx)")
    args = ap.parse_args()
    out = args.out or (os.path.splitext(args.file)[0] + "_tables.xlsx")
    try:
        n = convert(args.file, out)
    except H.NotHwpxError as e:
        print(str(e))
        return 2
    except FileNotFoundError:
        print(f"File not found: {args.file}")
        return 2
    print(f"Wrote {n} table sheet(s) → {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
