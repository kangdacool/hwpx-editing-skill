#!/usr/bin/env python3
"""
data_to_hwpx_table.py — insert an Excel/CSV table into an HWPX document, preserving
merged cells. The new table is appended to the end and styled by CLONING an
existing table in the target, so it reuses that document's valid cell definitions
(charPr/paraPr/borderFill), column width, and row height.

Usage:
    python data_to_hwpx_table.py DATA.(xlsx|csv) TARGET.hwpx [OUT.hwpx] [--sheet NAME]

- .xlsx : first sheet by default (or --sheet NAME); Excel merges → cell spans.
- .csv  : encoding auto-detected (utf-8 / cp949 / euc-kr); CSV has no merges.
- TARGET must contain at least one table (its first table is the style template).
  Open the result in 한글 to finalize (row heights auto-expand on open).

Requires: lxml; openpyxl only for .xlsx input.
"""
import argparse
import copy
import csv
import os
import sys
import zipfile

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError, OSError):
    pass

import hwpxlib as H

P = H.P


def read_data(path, sheet=None):
    """Return (grid: list[list[str]], merges: list[(r0,c0,rowSpan,colSpan)])."""
    if os.path.splitext(path)[1].lower() == ".csv":
        rows = None
        for enc in ("utf-8-sig", "cp949", "euc-kr"):
            try:
                with open(path, encoding=enc, newline="") as f:
                    rows = list(csv.reader(f))
                break
            except UnicodeDecodeError:
                continue
        if rows is None:
            raise SystemExit(f"Could not decode {path} as utf-8/cp949/euc-kr.")
        m = max((len(r) for r in rows), default=0)
        return [[(c or "") for c in r] + [""] * (m - len(r)) for r in rows], []
    try:
        import warnings
        warnings.filterwarnings("ignore")  # openpyxl "no default style" noise
        import openpyxl
    except ImportError as e:  # pragma: no cover
        raise SystemExit("Reading .xlsx needs openpyxl:  pip install openpyxl") from e
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    n, m = ws.max_row, ws.max_column

    def cv(v):
        if v is None:
            return ""
        return str(int(v)) if isinstance(v, float) and v.is_integer() else str(v)

    grid = [[cv(ws.cell(r + 1, c + 1).value) for c in range(m)] for r in range(n)]
    merges = [(mr.min_row - 1, mr.min_col - 1,
               mr.max_row - mr.min_row + 1, mr.max_col - mr.min_col + 1)
              for mr in ws.merged_cells.ranges]
    return grid, merges


def insert_table(target, grid, merges, out):
    """Append a table built from grid/merges to `target` (cloning its first table's
    cell style), write to `out`, and return (rows, cols, merge_count)."""
    H.ensure_hwpx(target)
    z = zipfile.ZipFile(target)
    sections = H.section_names(z)
    tpl_p = tpl_tc = None
    for name in sections:
        root0 = H.etree.fromstring(z.read(name))
        tbl0 = root0.find(f".//{P}tbl")
        if tbl0 is not None:
            tpl_p = next(a for a in tbl0.iterancestors()
                         if a.tag == f"{P}p" and a.getparent() is root0)
            tpl_tc = tbl0.find(f".//{P}tc")
            tblw = int(tbl0.find(f"{P}sz").get("width"))
            rowh = int(tpl_tc.find(f"{P}cellSz").get("height"))
            break
    if tpl_tc is None:
        raise SystemExit(
            "TARGET has no table to use as a cell-style template. 대상 문서에 표가 "
            "하나도 없습니다 — 한글에서 빈 표를 하나 넣고 저장한 뒤 다시 시도하거나, "
            "표가 있는 문서를 쓰세요.")

    last = sections[-1]
    root = H.etree.fromstring(z.read(last))
    n = len(grid)
    m = len(grid[0]) if grid else 0
    covered, spans = set(), {}
    for r0, c0, rs, cs in merges:
        spans[(r0, c0)] = (rs, cs)
        for r in range(r0, r0 + rs):
            for c in range(c0, c0 + cs):
                if (r, c) != (r0, c0):
                    covered.add((r, c))
    widths = [tblw // m] * m
    widths[-1] = tblw - sum(widths[:-1])  # cell widths sum exactly to table width

    uid = H.make_uid(root)
    new_p = copy.deepcopy(tpl_p)
    tbl = new_p.find(f".//{P}tbl")
    tbl.set("id", str(uid()))
    tbl.set("rowCnt", str(n))
    tbl.set("colCnt", str(m))
    tbl.find(f"{P}sz").set("height", str(rowh * n))
    for tr in tbl.findall(f"{P}tr"):
        tbl.remove(tr)
    for r in range(n):
        tr = H.etree.SubElement(tbl, f"{P}tr")
        for c in range(m):
            if (r, c) in covered:
                continue  # covered by a span → omit (HWPX convention)
            rs, cs = spans.get((r, c), (1, 1))
            tc = copy.deepcopy(tpl_tc)
            tc.find(f"{P}cellAddr").set("colAddr", str(c))
            tc.find(f"{P}cellAddr").set("rowAddr", str(r))
            tc.find(f"{P}cellSpan").set("colSpan", str(cs))
            tc.find(f"{P}cellSpan").set("rowSpan", str(rs))
            tc.find(f"{P}cellSz").set("width", str(sum(widths[c:c + cs])))
            tc.find(f"{P}cellSz").set("height", str(rowh * rs))  # rowSpan → summed height
            para = tc.find(f".//{P}p")
            para.set("id", "0")
            ls = para.find(f"{P}linesegarray")
            if ls is not None:
                para.remove(ls)
            run = para.find(f"{P}run")
            for t in run.findall(f"{P}t"):
                run.remove(t)
            H.etree.SubElement(run, f"{P}t").text = grid[r][c]
            tr.append(tc)
    root.append(new_p)
    H.strip_linesegarray(root)
    new_bytes = H.XML_DECL + H.etree.tostring(root, encoding="unicode").encode("utf-8")
    H.repack_preserve(target, {last: new_bytes}, out)
    return n, m, len(spans)


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Insert an Excel/CSV table into an HWPX (merged cells preserved).")
    ap.add_argument("data", help="the .xlsx or .csv to insert")
    ap.add_argument("target", help="the .hwpx to insert into (must contain a table)")
    ap.add_argument("out", nargs="?", help="output .hwpx (default: <target>_with_table.hwpx)")
    ap.add_argument("--sheet", help="xlsx sheet name (default: first sheet)")
    args = ap.parse_args()
    out = args.out or (os.path.splitext(args.target)[0] + "_with_table.hwpx")
    try:
        grid, merges = read_data(args.data, args.sheet)
        n, m, sp = insert_table(args.target, grid, merges, out)
    except H.NotHwpxError as e:
        print(str(e))
        return 2
    except FileNotFoundError as e:
        print(f"File not found: {e.filename}")
        return 2
    print(f"Inserted a {n}x{m} table ({sp} merged cell(s)) -> {out}")
    print("Open it in 한글 to finalize layout (row heights auto-expand on open).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
